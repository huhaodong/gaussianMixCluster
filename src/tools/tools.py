#!/usr/bin/python
# -*- coding: UTF-8 -×-

import xlrd
from openpyxl import Workbook
import numpy as np


def showHelp():
    print '''
      python main.py -k <int> -i <file path> -o <file path>
      -h 获取帮助
      -k 设置聚类的个数,默认为3个
      -i 指定输入文件
      -o 指定输出文件夹
      -l 指定最大迭代次数
      '''


def loadFile(fileName):
    data = xlrd.open_workbook(fileName)
    table = data.sheets()[0]
    first = table.col_values(0)
    second = table.col_values(1)
    tri = table.col_values(2)
    nrows = table.nrows
    dataSet = []
    result = []
    for i in range(nrows):
        dataSet.append([float(first[i]), float(second[i])])
        result.append(int(tri[i]))
    return (dataSet,result)


def writeFile(sourceData, classData, fileName):
    sourceMat = np.mat(sourceData)
    classMat = np.mat(classData)
    r1, c1 = sourceMat.shape
    r2, c2 = classMat.shape

    wb = Workbook()
    ws = wb.get_active_sheet()
    ws.title = 'output'
    # new_ws = wb.create_sheet(title='output')
    for i in range(r1):
        for j in range(c1):
            ws.cell(row=i + 1, column=j + 1).value = sourceMat[i, j]
    for i in range(r2):
        for j in range(c2):
            ws.cell(row=j + 1, column=i + 1 + c1).value = classMat[i, j]
            # worksheet.write(i,j+c1,classMat[i,j])
    wb.save(filename=fileName)
